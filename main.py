import constants
import shutil
import datetime
import openpyxl
import win32com.client as win32
from os import path
import docx

class paperwork:
	today = datetime.datetime.today()
	today_yyyymmdd = today.strftime('%Y%m%d')
	today_ddmonyyyy = today.strftime('%d-%b-%Y')
	
	def __init__(self):
		""" Initialise the variables """
		
		self.chrq_num = input("CHRQ number:");
		self.sf_num = input("SF number:")
		self.title = input("Title:")
		self.description = input("Description:")
		#self.description = self.get_multi_line_input("Description:")
		self.qc_num = input("QC number:")
		self.cr_num = input("CR number:")
		self.patch_num = input("Patch number:")
		self.bespoke = input("Bespoke change(Yy/Nn):").upper()
		self.expected_dev_effort = input("Expected development effort:")
		self.type_of_dev = input("Type of development (Bug fix, Enhancement, New Development):")
		self.code = {}
		
		'''
		self.chrq_num = '44444343'
		self.sf_num = '3333'
		self.title = 'Title for 4321'
		self.description = 'Description for 4321'
		self.patch_num = '22331'
		self.qc_num = 'QC for 4321'
		self.cr_num = 'CR for 4321'
		self.code = {'abc.sql':'11.22.33', 'abc.sh':'11.22.33','abc.rex':'11.22.33','abc.pc':'11.22.33','abc.sss':'11.22.33'}
		self.bespoke = 'N'
		self.expected_dev_effort = '4'
		self.type_of_dev = "Enhancement"
		'''
		
		while True:
			self.code_name = input("Code name:")
			if self.code_name.lower() == "end":
				break
			self.code_version = input("Code version:")
			self.code[self.code_name] = self.code_version;
		
	def get_multi_line_input(self, prompt):
		print(prompt+':')
		input_list = []

		while True:
			input_str = input(">")
			if input_str == "": #and input_list[-1] == "":
				break
			else:
				input_list.append(input_str)
		return "\n".join(input_list)
	
	def generate_utp(self):
		"""Generates UTP document"""
		
		#UTP file name
		utp_file_name = r'CR_'+ self.chrq_num + '_' + 'Unit_Test_Plan_' + constants.initials + '_' + self.today_yyyymmdd + '.xlsx'
		utp_file_temp = constants.temp_dir + utp_file_name
		
		#Copy base document
		shutil.copy(constants.utp_base_document,utp_file_temp)
		
		#Edit the UTP
		wb = openpyxl.load_workbook(filename = utp_file_temp)
		
		########DOCUMENT CONTROL TAB############# 
		ws = wb.get_sheet_by_name("Document Control")
		
		ws.cell(row=16,column=1).value='CHRQ ' + self.chrq_num + '-' + self.title;
		ws.cell(row=37,column=2).value=constants.developer_name
		ws.cell(row=38,column=2).value=constants.developer_name
		ws.cell(row=39,column=2).value='v1'
		ws.cell(row=40,column=2).value=self.today_ddmonyyyy
		ws.cell(row=42,column=2).value=constants.authoriser
		
		#######PLAN AND DEFINITION TAB###########
		ws = wb.get_sheet_by_name("Plan and Definition")
		
		#Test Design Details
		ws.cell(row=5,column=2).value=self.chrq_num
		ws.cell(row=6,column=2).value=self.title
		
		ws.cell(row=10,column=2).value=self.description
		ws.cell(row=10,column=2).style.alignment.wrap_text = True
		
		ws.cell(row=14,column=2).value=constants.test_environment
		ws.cell(row=15,column=2).value=constants.trga
		ws.cell(row=16,column=2).value=constants.developer_name
		
		#Code Under Test
		i=0
		for code_name in self.code:
			cur_row=32+i
			ws.cell(row=cur_row,column=1).value=code_name
			ws.cell(row=cur_row,column=3).value=self.code[code_name]
			i=i+1

		#############Test Plan##############
		ws = wb.get_sheet_by_name("Test Plan")
		
		start=6
		for  i in range(0,5):
			ws.cell(row=start+i,column=7).value='As Expected'
			ws.cell(row=start+i,column=8).value=constants.initials
			ws.cell(row=start+i,column=9).value=self.today_ddmonyyyy
			ws.cell(row=start+i,column=10).value='PASS'
		
		wb.save(utp_file_temp)
		
		#Copy to shared directory
		shutil.copy(utp_file_temp,constants.utp_path)
		
	def _get_file_dir(self, file_ext):
		return{
			'.sql':'$RIMS_SQL',
			'.sh':'$RIMS_COM',
			'.pc':'$RIMS_PC',
			'.cfg':'$RIMS_DATA',
			'.rex':'$RIMS_CON'}.get(file_ext,'$RIMS_')
		
	def generate_code_review(self):
		"""Generates code review document"""
		
		#Code review file name
		code_review_file_name = r'Code review - CHRQ - ' + self.chrq_num + ' - ' + self.title + ' - '+ self.today_ddmonyyyy + '.xlsx'
		code_review_file_temp = constants.temp_dir + code_review_file_name 
		
		#Copy base document
		shutil.copy(constants.code_review_base_document,code_review_file_name)
		
		#Edit code review document
		wb = openpyxl.load_workbook(filename = code_review_file_name)
		
		##########Code Review Tab##########
		ws = wb.get_sheet_by_name("Code Review")
		
		ws.cell(row=2, column=4).value=self.chrq_num+'/'+self.sf_num	#CHRQ/SF
		ws.cell(row=5, column=3).value=self.chrq_num					#CHRQ number
		ws.cell(row=5, column=4).value=self.title						#Description
		
		i=0
		for code_name in self.code:
			cur_row=10+i
			ws.cell(row=cur_row,column=4).value=code_name				#File name
			ws.cell(row=cur_row,column=5).value=self.code[code_name]	#File version 
			ws.cell(row=cur_row,column=6).value='1'						#Comment number
			ws.cell(row=cur_row,column=7).value='No change required'	#Comment
			ws.cell(row=cur_row,column=3).value=self._get_file_dir(path.splitext(code_name)[1])		#Directory
			i=i+1
		
		wb.save(code_review_file_temp)
		
		#Copy to shared directory
		shutil.copy(code_review_file_temp,constants.code_review_path)
		
	def generate_STR_form(self):
		"""Generates System Test Request form"""
		#STR file name
		str_file_name = 'CR'+self.chrq_num+'_'+constants.initials+'_'+self.today_yyyymmdd+'_v1.docx'
		str_file_temp = constants.temp_dir+str_file_name
		
		#Copy base document
		shutil.copy(constants.str_base_document,str_file_temp)
		
		#Edit code review file
		document = docx.Document(str_file_temp)
		
		#SystemRequestTestDetails(srtd)
		srtd = document.tables[0]
		
		srtd.cell(2,1).text = self.today_ddmonyyyy		#Date
		srtd.cell(3,1).text = '1'						#Version of submission
		srtd.cell(4,1).text = constants.developer_name+'/'+constants.email+'/'+constants.extension		#Submitted by(name / e-mail / telephone #)
		srtd.cell(5,1).text = self.title				#Title of development
		srtd.cell(6,1).text = constants.client_name		#Client Name
		srtd.cell(7,1).text = constants.billable_qa		#BillableQA (Y/N)
		srtd.cell(8,1).text = constants.development_manager		#Development Manager
		srtd.cell(9,1).text = constants.developer_name	#Contact Developers
		srtd.cell(10,1).text = self.description			#Short description
		srtd.cell(11,1).text = constants.system_name	#System Name
		srtd.cell(12,1).text = constants.system_version #System Version
		srtd.cell(13,1).text = (self.today + datetime.timedelta(days=int(constants.days_for_delivery_date))).strftime('%d-%b-%Y')  #Expected delivery date to QA
		srtd.cell(14,1).text = (self.today + datetime.timedelta(days=int(constants.days_for_delivery_date)+int(constants.days_for_qa_testing))).strftime('%d-%b-%Y') #Date requested to be tested By
		srtd.cell(15,1).text = self.expected_dev_effort #Expected development effort
		srtd.cell(16,1).text = constants.complexity		#Complexity
		srtd.cell(17,1).text = constants.priority		#Priority
		srtd.cell(18,1).text = self.description			#Detailed description of functionality
		srtd.cell(20,1).text = self.type_of_dev			#Type of development
		srtd.cell(21,1).text = self.chrq_num			#CHRQ Number and/or BT Num
		srtd.cell(23,1).text = self.sf_num				#Other Reference (type & number i.e. SupportForce)
		srtd.cell(24,1).text = r'CR_'+ self.chrq_num + '_' + 'Unit_Test_Plan_' + constants.initials + '_' + self.today_yyyymmdd + '.xlsx' #UTP file name
		srtd.cell(25,1).text = constants.utp_path		#Unit test plan location
		if self.bespoke == 'Y':
			srtd.cell(29,1).text = constants.special_note  #Special note
			
		document.save(str_file_temp)
		
		#Copy to shared directory
		shutil.copy(str_file_temp,constants.str_path)
		
class mailer:
	def __init__(self,paperwork_object):
		self.title = ''
		self.greet = ''
		self.body = ''
		self.to = ''
		self.cc = ''
		self.paperwork_object = paperwork_object
		
	def generate_mail(self,subject,body,to,cc):
		outlook = win32.Dispatch('outlook.application')
		mail = outlook.CreateItem(0)
		mail.To = to
		mail.Cc = cc
		mail.Subject = subject
		#mail.HtmlBody = body
		mail.Body = body
		mail.save()
		#mail.display(True)
		
	def chrq_approval_mail(self):
		self.subject = 'CHRQ ' + self.paperwork_object.chrq_num + '- '+self.paperwork_object.title + ' [SF '+ self.paperwork_object.sf_num +']'
		self.greet = constants.chrq_approval_greet
		body_line1 = 'CHRQ ' + self.paperwork_object.chrq_num + ' is created for '+ self.paperwork_object.title+'.'
		body_line2 = 'Could you please review and approve this change for me to proceed?'
		self.body = "\n".join([self.greet,"\n",body_line1,body_line2])
		self.to = constants.chrq_approval_to
		self.cc = constants.chrq_approval_cc
		self.generate_mail(self.subject,self.body,self.to,self.cc)
		
	def patch_mail(self):
		self.subject = '['+constants.client_name + ' '+constants.RIMS_version +'] - Patch Installtion Request -' + self.paperwork_object.patch_num + '- ' + self.paperwork_object.title + ' [SF - ' + self.paperwork_object.sf_num + ']'
		self.greet = constants.patch_greet
		body_line1 = 'Can you please install patch ' + self.paperwork_object.patch_num + ' in '+ constants.client_name + ' '+constants.RIMS_version+'?'
		body_line2 = 'The SF is ' + self.paperwork_object.sf_num+'.'
		self.body = "\n".join([self.greet,"\n",body_line1,body_line2])
		self.to = constants.patch_to
		self.cc = constants.patch_cc
		self.generate_mail(self.subject,self.body,self.to,self.cc)
				
	def com_request_mail(self):
		self.subject = 'CHRQ '+self.paperwork_object.chrq_num +' -' + self.paperwork_object.title+ '- [SF' + self.paperwork_object.sf_num+']'
		self.greet = constants.com_request_mail_greet
		body_line1 = 'This is w.r.t CHRQ ' + self.paperwork_object.chrq_num + ' - '+self.paperwork_object.title +'.'
		body_line2 = 'We have to deliver to '+constants.client_name+'.'
		body_line3 = 'Due to the bespoke nature of the change , we are requesting dispensation for the system test but will provide details of the testing conducted in the release area for QA validation (I have added this quote in the STR as well).'
		body_line4 = 'Can you please approve and COM this CHRQ?'
		body_line5 = 'The STR and UTP are stored at the designated location.'
		if self.paperwork_object.bespoke == "Y":
			self.body = "\n".join([self.greet,"\n",body_line1,body_line2,body_line3,body_line4,body_line5])
		else:
			self.body = "\n".join([self.greet,"\n",body_line1,body_line2,body_line4,body_line5])
		self.to = constants.com_request_mail_to
		self.cc = constants.com_request_mail_cc
		self.generate_mail(self.subject,self.body,self.to,self.cc)